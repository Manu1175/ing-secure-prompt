"""Excel workbook utilities for SecurePrompt."""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

from openpyxl import load_workbook

from secureprompt.entities.detectors import detect
from secureprompt.prompt.lexicon import to_tokens as lex_tokens
from secureprompt.receipts.store import write_receipt
from secureprompt.scrub import pipeline as scrub_pipeline
from secureprompt.ui.selective import selective_sanitize


CellKey = Tuple[str, str]


def read_xlsx_text(path: Path) -> List[Dict[str, Any]]:
    """Extract textual content from non-empty cells.

    Parameters
    ----------
    path:
        Absolute path to the workbook on disk.

    Returns
    -------
    list of dict
        One entry per populated cell containing ``sheet``, ``cell`` and ``text`` keys.
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    cells: List[Dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text:
                        continue
                    cells.append({"sheet": sheet.title, "cell": cell.coordinate, "text": text})
    finally:
        workbook.close()
    return cells


def write_xlsx_redacted(
    src_path: Path,
    replacements: Dict[CellKey, str],
    *,
    output_path: Optional[Path] = None,
) -> Path:
    """Write a redacted workbook next to the original.

    Parameters
    ----------
    src_path:
        Template workbook whose values will be replaced.
    replacements:
        Mapping from ``(sheet, cell)`` to sanitized text values.
    output_path:
        Optional destination path. Defaults to ``<stem>.redacted.xlsx`` besides ``src_path``.

    Returns
    -------
    Path
        Location of the new redacted workbook.
    """

    workbook = load_workbook(src_path)
    try:
        for (sheet_name, cell_ref), text in replacements.items():
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            worksheet[cell_ref].value = text

        destination = output_path or src_path.with_name(f"{src_path.stem}.redacted{src_path.suffix}")
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(destination))
    finally:
        workbook.close()

    return destination


def _ids_to_tokens(s: str) -> str:
    return re.sub(r"\bC[1-5]::([A-Z0-9_]+)::[0-9a-f]{8,40}\b", lambda m: f"<{m.group(1)}>", s)


def _scrub_cell_text(text: str, base_c_level: str = "C4") -> Tuple[str, str, List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Apply rule-based scrubbing to a single cell value."""

    hits = detect(text)
    transformed = text
    public_entities: List[Dict[str, Any]] = []
    receipt_entities: List[Dict[str, Any]] = []

    for hit in sorted(hits, key=lambda item: item["start"], reverse=True):
        label = hit["label"]
        entity_level = scrub_pipeline.ENTITY_DEFAULT_C_LEVEL.get(label, base_c_level)
        identifier = scrub_pipeline._identifier(label, hit["value"], entity_level)
        action = hit.get("action", "redact")
        mask_preview = scrub_pipeline._mask_value(hit["value"]) if action == "mask" else None

        transformed = transformed[: hit["start"]] + identifier + transformed[hit["end"] :]

        entity = {
            "label": label,
            "span": [hit["start"], hit["end"]],
            "detector": hit["rule_id"],
            "confidence": hit["confidence"],
            "c_level": entity_level,
            "identifier": identifier,
            "action": action,
        }
        if mask_preview is not None:
            entity["mask_preview"] = mask_preview
        public_entities.append(entity)

        receipt_entities.append(
            {
                "identifier": identifier,
                "label": label,
                "detector": hit["rule_id"],
                "c_level": entity_level,
                "confidence": hit["confidence"],
                "span": [hit["start"], hit["end"]],
                "original": hit["value"],
            }
        )

    public_entities.reverse()
    receipt_entities.reverse()

    tokenized_from_raw = lex_tokens(text or "")
    tokenized_from_ids = _ids_to_tokens(transformed)
    tokenized = tokenized_from_ids if tokenized_from_raw == (text or "") else tokenized_from_raw

    return transformed, tokenized, public_entities, receipt_entities


def scrub_workbook(path: Path, clearance: str, *, filename: Optional[str] = None) -> Dict[str, Any]:
    """Scrub an Excel workbook and persist a redacted artefact with receipt metadata.

    Parameters
    ----------
    path:
        Local filesystem path to the uploaded workbook.
    clearance:
        Clearance selected in the UI (C1–C4) used for selective sanitisation.
    filename:
        Optional original filename for logging purposes.

    Returns
    -------
    dict
        Structured information including operation id, receipt path, redacted file path,
        entities for UI display, combined hashes, and summary strings.
    """

    rows = read_xlsx_text(path)
    clearance = clearance.upper()

    combined_original_segments: List[str] = []
    combined_scrubbed_segments: List[str] = []
    display_sanitized_segments: List[str] = []
    placeholder_map: Dict[str, str] = {}
    replacements: Dict[CellKey, str] = {}
    entities_for_ui: List[Dict[str, Any]] = []
    receipt_entities: List[Dict[str, Any]] = []

    for row in rows:
        sheet = row["sheet"]
        cell_ref = row["cell"]
        cell_text = row["text"]

        _scrubbed_ids, scrubbed_tokens, cell_public_entities, cell_receipt_entities = _scrub_cell_text(cell_text, base_c_level="C4")

        # The persisted workbook should never contain raw sensitive values.
        replacements[(sheet, cell_ref)] = scrubbed_tokens
        combined_original_segments.append(f"{sheet}!{cell_ref}={cell_text}")
        combined_scrubbed_segments.append(f"{sheet}!{cell_ref}={scrubbed_tokens}")

        sanitized_for_display = selective_sanitize(cell_text, cell_public_entities, clearance)
        if sanitized_for_display == cell_text and scrubbed_tokens != cell_text:
            sanitized_for_display = scrubbed_tokens
        display_sanitized_segments.append(f"{sheet}!{cell_ref}={sanitized_for_display}")

        for public_entity, receipt_entity in zip(cell_public_entities, cell_receipt_entities):
            excel_meta = {"sheet": sheet, "cell": cell_ref, "offset": receipt_entity["span"]}

            entity_copy = dict(public_entity)
            entity_copy["excel"] = excel_meta
            entities_for_ui.append(entity_copy)

            receipt_copy = dict(receipt_entity)
            receipt_copy["excel"] = excel_meta
            receipt_entities.append(receipt_copy)

            placeholder_map[public_entity["identifier"]] = public_entity["identifier"]

    operation_id = uuid4().hex

    combined_original = "\n".join(combined_original_segments)
    combined_scrubbed = "\n".join(combined_scrubbed_segments)

    original_hash = hashlib.sha256(combined_original.encode("utf-8")).hexdigest()
    scrubbed_hash = hashlib.sha256(combined_scrubbed.encode("utf-8")).hexdigest()

    source_bytes = path.stat().st_size if path.exists() else 0

    receipt_path = write_receipt(
        operation_id=operation_id,
        text=combined_original,
        scrubbed=combined_scrubbed,
        entities=receipt_entities,
        c_level="C4",
        filename=filename,
        policy_version=None,
        placeholder_map=placeholder_map,
    )

    redacted_root = Path("data/redacted") / operation_id
    redacted_root.mkdir(parents=True, exist_ok=True)

    original_name = filename or path.name
    redacted_filename = f"{Path(original_name).stem}.redacted.xlsx"
    redacted_path = write_xlsx_redacted(
        src_path=path,
        replacements=replacements,
        output_path=redacted_root / redacted_filename,
    )

    # Remove the original temporary upload once the redacted artefact is produced.
    try:
        path.unlink()
    except FileNotFoundError:
        pass

    return {
        "operation_id": operation_id,
        "receipt_path": str(receipt_path),
        "redacted_path": str(redacted_path),
        "entities": entities_for_ui,
        "original_display": combined_original,
        "sanitized_display": "\n".join(display_sanitized_segments),
        "original_hash": original_hash,
        "scrubbed_hash": scrubbed_hash,
        "source_bytes": source_bytes,
        "scrubbed_text": combined_scrubbed,
    }


__all__ = ["read_xlsx_text", "write_xlsx_redacted", "scrub_workbook"]
