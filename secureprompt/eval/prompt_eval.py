from __future__ import annotations

import json
import os
import re
import difflib
from collections import Counter
from pathlib import Path
from collections.abc import Iterable
from typing import Any, Dict, List, Optional, Tuple
from urllib import error, request

from openpyxl import Workbook
from openpyxl import load_workbook

from secureprompt.prompt.sanitizer import sanitize_prompt as SAN

_SCRUB_WARNING_SHOWN = False
_SCRUB_URL = os.getenv("SP_SCRUB_URL", "http://127.0.0.1:8000/scrub")

_EXPECTED_COLUMNS = {
    "original_prompt": ("original", "prompt"),
    "expected_sanitized_prompt": ("sanitized", "prompt"),
    "response": ("response",),
    "expected_sanitized_response": ("sanitized", "response"),
}

_APPENDED_COLUMNS = [
    "Got_Sanitized_Prompt",
    "Got_Sanitized_Response",
    "Prompt_Replacements",
    "Response_Entities",
    "Prompt_Correct",
    "Response_Correct",
    "Receipt_Path",
]

TOKEN_RX = re.compile(r"[<\[]([A-Z0-9_]+)[>\]]")
_SMART_QUOTES = {
    "“": '"',
    "”": '"',
    "„": '"',
    "‟": '"',
    "‘": "'",
    "’": "'",
    "‚": "'",
    "‛": "'",
}


def detect_clearance(path: str | Path, default: str = "C3") -> str:
    """Infer the clearance level from env or filename (fallbacks to default)."""

    default = (default or "C3").upper()
    choices = {"C1", "C2", "C3", "C4"}
    env = os.getenv("SP_CLEARANCE")
    if env:
        env = env.strip().upper()
        if env in choices:
            return env
    basename = Path(path).name.lower()
    match = re.search(r"c([1-4])", basename)
    if match:
        candidate = f"C{match.group(1)}"
        if candidate in choices:
            return candidate
    return default if default in choices else "C3"


def _normalize(text: Any) -> str:
    return str(text).strip() if isinstance(text, str) else (str(text).strip() if text is not None else "")


def _normalize_header(name: str) -> Tuple[str, ...]:
    cleaned = re.sub(r"[^a-z0-9]+", " ", (name or "").lower()).split()
    return tuple(cleaned)


def _match_column(name: str) -> Optional[str]:
    tokens = set(_normalize_header(name))
    if not tokens:
        return None
    for key, required in _EXPECTED_COLUMNS.items():
        if all(token in tokens for token in required):
            return key
    return None


def read_sheet(path: str | Path) -> List[Dict[str, Any]]:
    """Read the first sheet of an XLSX workbook into normalized row dicts."""

    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook.active
    headers: List[str] = []
    header_map: Dict[int, str] = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []
    for idx, cell in enumerate(header_row, start=1):
        header = _normalize(cell)
        headers.append(header)
        mapped = _match_column(header)
        if mapped:
            header_map[idx] = mapped
    header_values = set(header_map.values())
    has_expected_prompt = "expected_sanitized_prompt" in header_values
    has_expected_response = "expected_sanitized_response" in header_values
    rows: List[Dict[str, Any]] = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_row: Dict[str, Any] = {}
        normalized: Dict[str, Any] = {
            "__row_index__": row_idx,
            "__headers__": headers,
            "__raw__": raw_row,
        }
        for col_idx, cell_value in enumerate(values, start=1):
            header = headers[col_idx - 1] if col_idx - 1 < len(headers) else f"Column {col_idx}"
            raw_row[header] = cell_value
            mapped_key = header_map.get(col_idx)
            if mapped_key:
                normalized[mapped_key] = _normalize(cell_value)
        for required in _EXPECTED_COLUMNS:
            normalized.setdefault(required, "")
        normalized["__has_expected_prompt__"] = has_expected_prompt
        normalized["__has_expected_response__"] = has_expected_response
        rows.append(normalized)
    return rows


def _sanitize_prompt(
    text: str,
    xlsx_hint: str | Path | None = None,
    *,
    style: str = "square",
) -> Tuple[str, List[Any]]:
    if not text:
        return "", []
    result = SAN(text, xlsx_hint=xlsx_hint, style=style)
    if isinstance(result, tuple) and len(result) >= 2:
        sanitized_text, operations = result[0], result[1]
    elif isinstance(result, dict):  # pragma: no cover - defensive branch
        sanitized_text = result.get("text") or result.get("sanitized") or text
        operations = result.get("operations") or result.get("ops") or result.get("replacements") or []
    else:  # pragma: no cover - unexpected return type
        sanitized_text, operations = result, []
    if isinstance(operations, dict):
        replacements = operations.get("replacements")
        if isinstance(replacements, Iterable) and not isinstance(replacements, (str, bytes)):
            operations = replacements
        else:
            operations = list(operations.values())
    if isinstance(operations, (str, bytes)):
        operations_list: List[Any] = []
    elif isinstance(operations, Iterable):
        operations_list = list(operations)
    else:
        operations_list = []
    return str(sanitized_text), operations_list


def _count_operations(operations: Iterable[Any]) -> int:
    count = 0
    for _ in operations:
        count += 1
    return count


def _post_scrub(text: str, clearance: str) -> Dict[str, Any]:
    global _SCRUB_WARNING_SHOWN
    payload = json.dumps({"text": text, "c_level": clearance}).encode("utf-8")
    req = request.Request(_SCRUB_URL, data=payload, headers={"Content-Type": "application/json"})
    try:
        with request.urlopen(req, timeout=10) as response:
            body = response.read().decode("utf-8")
        return json.loads(body) if body else {}
    except (error.URLError, error.HTTPError, TimeoutError, json.JSONDecodeError, OSError):
        if not _SCRUB_WARNING_SHOWN:
            print("warning: scrub API unreachable; using raw responses")
            _SCRUB_WARNING_SHOWN = True
        return {
            "sanitized": text,
            "findings": [],
        }


def _extract_sanitized_text(data: Dict[str, Any], fallback: str) -> str:
    for key in ("sanitized_text", "sanitized", "output", "text_sanitized"):
        value = data.get(key)
        if isinstance(value, str):
            return value
    result = data.get("result")
    if isinstance(result, dict):
        value = result.get("text")
        if isinstance(value, str):
            return value
    return fallback


def _extract_findings(data: Dict[str, Any]) -> Optional[List[Any]]:
    value = data.get("findings")
    if isinstance(value, list):
        return value
    result = data.get("result")
    if isinstance(result, dict):
        nested = result.get("findings")
        if isinstance(nested, list):
            return nested
    return None


def _extract_receipt_path(data: Dict[str, Any]) -> str:
    for key in ("receipt", "receipt_path"):
        value = data.get(key)
        if isinstance(value, str):
            return value
    result = data.get("result")
    if isinstance(result, dict):
        value = result.get("receipt") or result.get("receipt_path")
        if isinstance(value, str):
            return value
    return ""


def _count_diffs(original: str, sanitized: str) -> int:
    matcher = difflib.SequenceMatcher(None, original or "", sanitized or "")
    return sum(1 for tag, *_ in matcher.get_opcodes() if tag != "equal")


def eval_row(
    row: Dict[str, Any],
    clearance: str,
    xlsx_hint: str | Path | None = None,
    *,
    style: str = "square",
) -> Dict[str, Any]:
    original_prompt = row.get("original_prompt", "")
    expected_prompt = row.get("expected_sanitized_prompt") if row.get("__has_expected_prompt__") else None
    prompt_text, operations = _sanitize_prompt(
        original_prompt,
        xlsx_hint=xlsx_hint,
        style=style,
    )
    prompt_replacements = _count_operations(operations)
    prompt_ops = [getattr(op, "label", "") for op in operations]

    response_text = row.get("response", "")
    scrub_result = _post_scrub(response_text, clearance)
    sanitized_response = _extract_sanitized_text(scrub_result, response_text)
    findings = _extract_findings(scrub_result)
    response_entities = len(findings) if findings is not None else _count_diffs(response_text, sanitized_response)
    expected_response = row.get("expected_sanitized_response") if row.get("__has_expected_response__") else None

    prompt_correct = (prompt_text == expected_prompt) if expected_prompt is not None else None
    response_correct = (sanitized_response == expected_response) if expected_response is not None else None

    return {
        "row_index": row.get("__row_index__"),
        "got_sanitized_prompt": prompt_text,
        "got_sanitized_response": sanitized_response,
        "prompt_replacements": prompt_replacements,
        "response_entities": response_entities,
        "prompt_correct": prompt_correct,
        "response_correct": response_correct,
        "receipt_path": _extract_receipt_path(scrub_result),
        "expected_prompt": expected_prompt,
        "expected_response": expected_response,
        "original_prompt": original_prompt,
        "response": response_text,
        "prompt_ops": [label for label in prompt_ops if label],
    }


def summarize(
    results: Iterable[Dict[str, Any]],
    *,
    seen_tokens: Optional[Dict[str, int]] = None,
    matched_tokens: Optional[Dict[str, int]] = None,
) -> Dict[str, Any]:
    total_rows = 0
    prompt_evaluated = prompt_correct_count = 0
    response_evaluated = response_correct_count = 0
    total_prompt_replacements = 0
    total_response_entities = 0
    for result in results:
        total_rows += 1
        total_prompt_replacements += int(result.get("prompt_replacements", 0))
        total_response_entities += int(result.get("response_entities", 0))
        prompt_correct = result.get("prompt_correct")
        if prompt_correct is not None:
            prompt_evaluated += 1
            if prompt_correct:
                prompt_correct_count += 1
        response_correct = result.get("response_correct")
        if response_correct is not None:
            response_evaluated += 1
            if response_correct:
                response_correct_count += 1
    prompt_accuracy = (prompt_correct_count / prompt_evaluated) if prompt_evaluated else None
    response_accuracy = (response_correct_count / response_evaluated) if response_evaluated else None
    seen_tokens = {k: int(v) for k, v in (seen_tokens or {}).items()}
    matched_tokens = {k: int(v) for k, v in (matched_tokens or {}).items()}
    expected_token_list = sorted(seen_tokens)
    missed_tokens = sorted(
        (token for token in expected_token_list if matched_tokens.get(token, 0) == 0),
        key=lambda t: (-seen_tokens.get(t, 0), t),
    )
    token_stats = [
        {
            "token": token,
            "seen": seen_tokens.get(token, 0),
            "matched": matched_tokens.get(token, 0),
        }
        for token in sorted(
            seen_tokens,
            key=lambda t: (-seen_tokens[t], t),
        )
    ]

    return {
        "total_rows": total_rows,
        "prompt_evaluated": prompt_evaluated,
        "prompt_correct": prompt_correct_count,
        "prompt_accuracy": prompt_accuracy,
        "response_evaluated": response_evaluated,
        "response_correct": response_correct_count,
        "response_accuracy": response_accuracy,
        "total_prompt_replacements": total_prompt_replacements,
        "total_response_entities": total_response_entities,
        "ops_by_label": {},
        "expected_tokens": expected_token_list,
        "seen_tokens": dict(sorted(seen_tokens.items())),
        "matched_tokens": dict(sorted(matched_tokens.items())),
        "top_missed_tokens": missed_tokens[:20],
        "token_stats": token_stats,
    }


def _heuristic_label(expected: Optional[str], got: Optional[str]) -> str:
    expected = expected or ""
    got = got or ""
    if "PRODUCT_NAME" in expected and "PRODUCT_NAME" not in got:
        return "missed PRODUCT_NAME"
    if "[NAME" in expected and "[NAME" not in got:
        return "name not detected"
    if "POLICY" in expected.upper():
        return "policy pattern"
    return "mismatch"


def _short_diff(expected: Optional[str], got: Optional[str]) -> str:
    expected = expected or ""
    got = got or ""
    matcher = difflib.SequenceMatcher(None, expected, got)
    snippets: List[str] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            continue
        left = expected[i1:i2]
        right = got[j1:j2]
        if len(left) > 40:
            left = left[:37] + "..."
        if len(right) > 40:
            right = right[:37] + "..."
        snippets.append(f"{tag}: '{left}' -> '{right}'")
        if len(snippets) >= 3:
            break
    return "; ".join(snippets) or "differences detected"


def tokens_of(s: str | None) -> set[str]:
    return set(TOKEN_RX.findall(s or ""))


def _prepare_display(text: Optional[str]) -> str:
    value = (text or "").replace("\r", " ").replace("\n", " ")
    for src, dst in _SMART_QUOTES.items():
        value = value.replace(src, dst)
    value = TOKEN_RX.sub(lambda m: f"[{m.group(1)}]", value)
    value = re.sub(r"\s+", " ", value).strip()
    if len(value) > 120:
        value = value[:117] + "..."
    return value


def collect_token_coverage(
    rows: Iterable[Dict[str, Any]],
    results: Iterable[Dict[str, Any]],
) -> Tuple[Dict[str, int], Dict[str, int]]:
    seen: Counter[str] = Counter()
    matched: Counter[str] = Counter()
    result_map = {
        result.get("row_index"): result for result in results
    }
    for row in rows:
        row_idx = row.get("__row_index__")
        if not row.get("__has_expected_prompt__"):
            continue
        expected_text = row.get("expected_sanitized_prompt") or ""
        if not isinstance(expected_text, str):
            continue
        expected_tokens = tokens_of(expected_text)
        for token in expected_tokens:
            seen[token] += 1
        result = result_map.get(row_idx)
        got_tokens: set[str] = set()
        if result:
            got_text = result.get("got_sanitized_prompt") or ""
            if isinstance(got_text, str):
                got_tokens = tokens_of(got_text)
        for token in expected_tokens & got_tokens:
            matched[token] += 1
    return dict(seen), dict(matched)


def write_outputs(
    input_path: str | Path,
    rows: List[Dict[str, Any]],
    results: List[Dict[str, Any]],
    summary: Dict[str, Any],
    output_dir: str | Path | None = None,
) -> Dict[str, Path]:
    reports_dir = Path(output_dir) if output_dir is not None else Path("reports")
    reports_dir.mkdir(parents=True, exist_ok=True)

    base = Path(input_path).stem
    workbook_path = reports_dir / f"{base}_eval.xlsx"
    summary_path = reports_dir / "prompt_eval_summary.json"
    anomalies_path = reports_dir / "prompt_eval_anomalies.md"

    headers: List[str] = rows[0].get("__headers__", []) if rows else []

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation"
    ws.append(headers + _APPENDED_COLUMNS)
    if len(rows) != len(results):
        raise ValueError("Row/result length mismatch")
    for row, result in zip(rows, results):
        raw = row.get("__raw__", {})
        values = [raw.get(header, "") for header in headers]
        values.extend(
            [
                result.get("got_sanitized_prompt", ""),
                result.get("got_sanitized_response", ""),
                result.get("prompt_replacements", 0),
                result.get("response_entities", 0),
                result.get("prompt_correct"),
                result.get("response_correct"),
                result.get("receipt_path", ""),
            ]
        )
        ws.append(values)
    wb.save(workbook_path)

    with summary_path.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2, sort_keys=True)

    anomalies: List[str] = ["# Prompt Evaluation Anomalies", ""]
    token_stats = summary.get("token_stats") or []
    if token_stats:
        anomalies.append("| Token | Seen in gold | Matched |")
        anomalies.append("|-------|--------------|---------|")
        for entry in token_stats[:20]:
            anomalies.append(
                f"| {entry['token']} | {entry['seen']} | {entry['matched']} |"
            )
        anomalies.append("")
    missed_tokens = summary.get("top_missed_tokens") or []
    if missed_tokens:
        subset = ", ".join(missed_tokens[:10])
        anomalies.append(f"Coverage gaps: {subset}")
        anomalies.append("")
    mismatch_entries = []
    for result in results:
        row_id = result.get("row_index")
        expected_prompt = result.get("expected_prompt")
        prompt_mismatch = (
            expected_prompt is not None and result.get("got_sanitized_prompt") != expected_prompt
        )
        if prompt_mismatch:
            mismatch_entries.append(
                (
                    row_id,
                    "prompt",
                    _heuristic_label(expected_prompt, result.get("got_sanitized_prompt")),
                    result.get("got_sanitized_prompt"),
                    expected_prompt,
                )
            )
        expected_response = result.get("expected_response")
        response_mismatch = (
            expected_response is not None and result.get("got_sanitized_response") != expected_response
        )
        if response_mismatch:
            mismatch_entries.append(
                (
                    row_id,
                    "response",
                    _heuristic_label(expected_response, result.get("got_sanitized_response")),
                    result.get("got_sanitized_response"),
                    expected_response,
                )
            )
    if not mismatch_entries:
        anomalies.append("No anomalies detected.")
    else:
        for row_id, kind, label, got_value, expected_value in mismatch_entries:
            got_display = _prepare_display(got_value)
            expected_display = _prepare_display(expected_value)
            anomalies.append(
                f"- Row {row_id} ({kind}): {label} — GOT: {got_display} vs EXPECTED: {expected_display}"
            )
    anomalies.append("")
    with anomalies_path.open("w", encoding="utf-8") as handle:
        handle.write("\n".join(anomalies))

    return {
        "workbook": workbook_path,
        "summary": summary_path,
        "anomalies": anomalies_path,
    }


__all__ = [
    "detect_clearance",
    "read_sheet",
    "eval_row",
    "summarize",
    "write_outputs",
    "collect_expected_tokens",
    "collect_ops_by_label",
]
