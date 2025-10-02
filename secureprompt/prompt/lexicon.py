from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path
from typing import Dict, Iterable, Optional, Set, Tuple

from openpyxl import load_workbook

try:  # pragma: no cover - optional dependency
    import yaml  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    yaml = None  # type: ignore

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_TOKEN_RX = re.compile(r"<([A-Z0-9_]+)>")


def _resolve_path(path: str | Path) -> Path:
    candidate = Path(path)
    if candidate.is_absolute():
        return candidate
    absolute = (Path.cwd() / candidate).resolve()
    if absolute.exists():
        return absolute
    return (_PROJECT_ROOT / candidate).resolve()


def _load_yaml_dict(path: Path) -> Dict[str, str]:
    if not path.is_file():
        return {}
    try:
        text = path.read_text(encoding="utf-8")
    except OSError:
        return {}
    if yaml is not None:
        try:
            data = yaml.safe_load(text) or {}
        except Exception:
            return {}
        placeholders = data.get("placeholders") if isinstance(data, dict) else None
        if isinstance(placeholders, dict):
            return {
                str(label): str(pattern)
                for label, pattern in placeholders.items()
                if isinstance(label, str) and isinstance(pattern, str)
            }
        return {}
    # Minimal fallback parser for simple placeholder mappings
    result: Dict[str, str] = {}
    current = None
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if not raw_line.startswith((" ", "\t")):
            current = line.rstrip(":")
            continue
        if current != "placeholders":
            continue
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        try:
            decoded = bytes(value, "utf-8").decode("unicode_escape")
        except Exception:
            decoded = value
        result[key] = decoded
    return result


def _dump_yaml_dict(path: Path, mapping: Dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if yaml is not None:
        data = {"placeholders": OrderedDict(sorted(mapping.items()))}
        try:
            path.write_text(
                yaml.safe_dump(
                    data,
                    sort_keys=False,
                    default_flow_style=False,
                ),
                encoding="utf-8",
            )
            return
        except Exception:
            pass
    if not mapping:
        path.write_text("placeholders: {}\n", encoding="utf-8")
        return
    lines = ["placeholders:"]
    for label, pattern in sorted(mapping.items()):
        escaped = pattern.replace("\\", "\\\\").replace('"', '\\"')
        lines.append(f'  {label}: "{escaped}"')
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def load_static_yaml(path: str | Path = "config/prompt_lexicon.yml") -> Dict[str, str]:
    return _load_yaml_dict(_resolve_path(path))


def extract_tokens_from_workbook(xlsx_path: str | Path) -> Set[str]:
    path = _resolve_path(xlsx_path)
    if not path.is_file():
        return set()
    try:
        workbook = load_workbook(filename=path, data_only=True, read_only=True)
    except Exception:
        return set()
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return set()

    def _normalize(cell: Optional[str]) -> str:
        if cell is None:
            return ""
        return re.sub(r"[^a-z0-9]+", " ", str(cell).lower()).strip()

    sanitized_columns = {
        idx
        for idx, cell in enumerate(header_row)
        if cell is not None
        and "sanitized" in _normalize(cell)
        and "prompt" in _normalize(cell)
    }
    if not sanitized_columns:
        return set()

    tokens: Set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        for idx in sanitized_columns:
            if idx >= len(row):
                continue
            value = row[idx]
            if not isinstance(value, str):
                continue
            for match in _TOKEN_RX.findall(value):
                tokens.add(match.upper())
    return tokens


def heuristics_for(token: str) -> Optional[str]:
    name = token.upper()
    if name.endswith("_NAME") or name in {"OWNER_NAME", "PAYER_NAME", "PAYEE_NAME"}:
        return r"\b([A-Z][a-z]+(?:\s[A-Z][a-z]+){0,3})\b"
    if name.endswith("_DATE") or name in {"LAST_EXECUTED_DATE", "REVIEW_DATE"}:
        return r"\b(\d{4}-\d{2}-\d{2}|\d{1,2}\s?[A-Za-z]{3,9}\s?\d{2,4})\b"
    if name.endswith("_NUMBER") or name.endswith("_ID"):
        return r"\b([A-Z0-9][A-Z0-9\-\_/]{3,})\b"
    if re.fullmatch(r"YEAR(?:_\d+)?", name):
        return r"\b(19|20)\d{2}\b"
    if name == "IBAN":
        return r"\b[A-Z]{2}\d{2}[A-Z0-9]{10,30}\b"
    if name == "EMAIL":
        return r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b"
    if name in {"APPLICATION_NAME", "APP_NAME"} or (name.startswith("APP") and name.endswith("_NAME")):
        return r"\bApp[_\s]?\d{1,4}\b"
    if name == "PROVIDER":
        return r"\b(AWS|Azure|GCP|Google\s+Cloud|On[- ]?Prem)\b"
    if name == "ENVIRONMENT":
        return r"\b(Production|Acceptance|UAT|Development|Dev|Test|Prod)\b"
    if name == "TRANSACTION_TYPE":
        return r"\b(online|sepa|wire|transfer|direct\s*debit|card)\b"
    if name == "SERVICE_CHANNEL":
        return r"\b(web|mobile|branch|call[ \-]?center|atm)\b"
    if name == "AGREEMENT_TYPE":
        return r"\b(supplier|customer|vendor)\s+agreements?\b"
    if name == "DOCUMENT_LINK":
        return r"https?://[^\s>]+"
    if name == "STORAGE_CAPACITY":
        return r"\b\d+(?:\.\d+)?\s?(GB|MB|TB|GiB|MiB|TiB)\b"
    if name == "VM_NAME":
        return r"\bVM[_-]?[A-Za-z0-9_-]+\b"
    if name == "COST_MODEL":
        return r"\b(OpEx|CapEx|cost\s*model|chargeback|showback)\b"
    if name == "ENTITY_TYPE":
        return r"\b(entity|supplier|customer|vendor|partner|application|service)\b"
    return None


def build_autolex(tokens: Iterable[str]) -> Dict[str, str]:
    autolex: Dict[str, str] = {}
    for token in sorted({t.upper() for t in tokens}):
        pattern = heuristics_for(token)
        if pattern:
            autolex[token] = pattern
    return autolex


def ensure_autolex_from_workbook(
    xlsx_path: str | Path,
    out: str | Path = "config/prompt_lexicon.auto.yml",
) -> Tuple[Path, int]:
    tokens = extract_tokens_from_workbook(xlsx_path)
    mapping = build_autolex(tokens)
    output_path = _resolve_path(out)
    _dump_yaml_dict(output_path, mapping)
    return output_path, len(mapping)


__all__ = [
    "load_static_yaml",
    "extract_tokens_from_workbook",
    "heuristics_for",
    "build_autolex",
    "ensure_autolex_from_workbook",
]
