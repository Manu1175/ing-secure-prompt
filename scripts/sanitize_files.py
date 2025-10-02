#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import pathlib
from pathlib import Path
from typing import Iterable, Optional
from urllib import error, request

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from secureprompt.eval import prompt_eval

try:
    from PyPDF2 import PdfReader  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None

SCRUB_URL = os.getenv("SP_SCRUB_URL", "http://127.0.0.1:8000/scrub")


ROOT_DIR = Path(__file__).resolve().parents[1]


def call_scrub(text: str, clearance: str) -> tuple[str, int, str]:
    payload = json.dumps({"text": text, "c_level": clearance}).encode("utf-8")
    req = request.Request(SCRUB_URL, data=payload, headers={"Content-Type": "application/json"})
    try:
        with request.urlopen(req, timeout=15) as response:
            body = response.read().decode("utf-8")
        data = json.loads(body) if body else {}
    except (error.URLError, error.HTTPError, TimeoutError, json.JSONDecodeError, OSError):
        print("warning: scrub API unreachable; returning original text")
        return text, 0, ""

    sanitized = _extract_sanitized(data, text)
    findings = _extract_findings(data)
    receipt = _extract_receipt(data)
    return sanitized, len(findings) if findings is not None else 0, receipt


def _extract_sanitized(payload: dict, fallback: str) -> str:
    for key in ("sanitized_text", "sanitized", "output", "text_sanitized"):
        value = payload.get(key)
        if isinstance(value, str):
            return value
    result = payload.get("result")
    if isinstance(result, dict):
        value = result.get("text")
        if isinstance(value, str):
            return value
    return fallback


def _extract_findings(payload: dict) -> Optional[list]:
    value = payload.get("findings")
    if isinstance(value, list):
        return value
    result = payload.get("result")
    if isinstance(result, dict):
        nested = result.get("findings")
        if isinstance(nested, list):
            return nested
    return None


def _extract_receipt(payload: dict) -> str:
    for key in ("receipt", "receipt_path"):
        value = payload.get(key)
        if isinstance(value, str):
            return value
    result = payload.get("result")
    if isinstance(result, dict):
        value = result.get("receipt") or result.get("receipt_path")
        if isinstance(value, str):
            return value
    return ""


def infer_clearance(path: Path, override: Optional[str]) -> str:
    if override:
        return override.upper()
    env = os.getenv("SP_CLEARANCE")
    if env:
        return env.upper()
    name = path.name.lower()
    tokens = name.replace("-", "_")
    for level in ("C1", "C2", "C3", "C4"):
        marker = f"_{level.lower()}_"
        if marker in tokens:
            return level
        if f"-{level.lower()}-" in name:
            return level
    return prompt_eval.detect_clearance(path)


def iter_files(base: Path) -> Iterable[Path]:
    if base.is_file():
        yield base
        return
    for path in sorted(base.rglob("*")):
        if path.is_file():
            yield path


def sanitize_xlsx(path: Path, clearance: str, reports_dir: Path) -> tuple[Path, str]:
    workbook = load_workbook(filename=path, data_only=False)
    receipts: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip():
                    sanitized, _count, receipt = call_scrub(value, clearance)
                    cell.value = sanitized
                    if receipt:
                        receipts.append(receipt)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"{path.stem}_sanitized.xlsx"
    workbook.save(out_path)
    receipt = receipts[-1] if receipts else ""
    return out_path, receipt


def sanitize_text_file(path: Path, clearance: str, reports_dir: Path) -> tuple[Path, str]:
    text = path.read_text(encoding="utf-8")
    sanitized, _count, receipt = call_scrub(text, clearance)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"{path.stem}_sanitized{path.suffix}"
    out_path.write_text(sanitized, encoding="utf-8")
    return out_path, receipt


def sanitize_pdf(path: Path, clearance: str, reports_dir: Path) -> Optional[tuple[Path, str]]:
    if PdfReader is None:
        print(f"warning: PyPDF2 not available; skipping {path}")
        return None
    try:
        reader = PdfReader(str(path))
    except Exception:
        print(f"warning: failed to read {path}; skipping")
        return None
    text_segments = []
    for page in reader.pages:
        try:
            segment = page.extract_text() or ""
        except Exception:
            segment = ""
        if segment:
            text_segments.append(segment)
    if not text_segments:
        print(f"warning: no extractable text in {path}; skipping")
        return None
    full_text = "\n".join(text_segments)
    sanitized, _count, receipt = call_scrub(full_text, clearance)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"{path.stem}_sanitized.txt"
    out_path.write_text(sanitized, encoding="utf-8")
    return out_path, receipt


def main() -> int:
    parser = argparse.ArgumentParser(description="Sanitize sample business files using the scrub API.")
    parser.add_argument("--in", dest="input_path", default="DATA", help="Path to a file or directory (default: DATA)")
    parser.add_argument("--clearance", dest="clearance", help="Clearance override (C1..C4)")
    args = parser.parse_args()

    target = Path(args.input_path)
    if not target.exists():
        alt = ROOT_DIR / target
        if not alt.exists():
            raise FileNotFoundError(f"Path not found: {target}")
        target = alt

    reports_dir = Path("reports")
    reports_dir.mkdir(parents=True, exist_ok=True)
    index_path = reports_dir / "sanitized_index.csv"
    entries: list[tuple[str, str, str]] = []

    for file_path in iter_files(target):
        suffix = file_path.suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".txt", ".html", ".htm", ".pdf"}:
            continue
        clearance = infer_clearance(file_path, args.clearance)
        sanitized_entry: Optional[tuple[Path, str]] = None
        if suffix in {".xlsx", ".xlsm"}:
            sanitized_entry = sanitize_xlsx(file_path, clearance, reports_dir)
        elif suffix in {".txt", ".html", ".htm"}:
            sanitized_entry = sanitize_text_file(file_path, clearance, reports_dir)
        elif suffix == ".pdf":
            sanitized_entry = sanitize_pdf(file_path, clearance, reports_dir)
        if sanitized_entry is None:
            continue
        sanitized_path, receipt = sanitized_entry
        entries.append((str(file_path), str(sanitized_path), receipt))

    if entries:
        with index_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["source", "sanitized", "receipt"])
            writer.writerows(entries)
    total_files = len(entries)
    receipt_count = sum(1 for _src, _dst, receipt in entries if receipt)
    print(f"Files: {total_files} | Receipts: {receipt_count} | Out: {index_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
