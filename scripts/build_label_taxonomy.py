from __future__ import annotations
import argparse, os, re, json, collections, glob, yaml
from openpyxl import load_workbook

# Accept both <FOO> and [FOO]
TOKEN_RX = re.compile(r"[<\[]([A-Z0-9_]+)[>\]]")
DROP_TRAILING_NUM = re.compile(r"(?:_\d+|[0-9])+$")

DEFAULT_C_LEVEL = "C2"
CLEVEL_GUESS = {
    "PAN":"C4","IBAN":"C4","BIC":"C3","EMAIL":"C3","PHONE":"C3","IPV4":"C3","IPV6":"C3",
    "NATIONAL_ID":"C4","DOB":"C3","NAME":"C2","ADDRESS":"C3","ACCOUNT_ID":"C3",
}

def norm(tok:str)->str:
    t = tok.strip().upper()
    t = re.sub(r"__+","_",t)
    return DROP_TRAILING_NUM.sub("", t)

def tokens_of(s:str|None):
    return [] if not s else [norm(m.group(1)) for m in TOKEN_RX.finditer(s)]

def scan_xlsx(path:str, counts:collections.Counter):
    wb = load_workbook(path, read_only=True); ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name:i for i,name in enumerate(hdr)}
    for key in ("Sanitized Prompt","Sanitized Response"):
        if key not in idx: continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            for t in tokens_of(row[idx[key]]):
                counts[t]+=1

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--prompts", default="PROMPTS", help="Folder with .xlsx prompt workbooks")
    ap.add_argument("--out", default="config/label_taxonomy.yml")
    args = ap.parse_args()

    counts = collections.Counter()
    for p in glob.glob(os.path.join(args.prompts, "*.xlsx")):
        scan_xlsx(p, counts)

    labels = sorted(counts.keys())
    manifest = {
        "generated_from": args.prompts,
        "total_labels": len(labels),
        "labels": [
            {"label": l, "c_level": CLEVEL_GUESS.get(l, DEFAULT_C_LEVEL), "examples": counts[l]}
            for l in labels
        ],
    }
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        yaml.safe_dump(manifest, f, sort_keys=False)
    print(json.dumps({"labels": len(labels), "out": args.out}, indent=2))

if __name__ == "__main__":
    main()
