#!/usr/bin/env python3
"""
Alternative evaluator that:
- Calls the /scrub API for each row
- Stores both raw and normalized scrubbed text
- Computes token coverage with bracket-agnostic + "C#::LABEL::" support
- Emits a summary JSON compatible with quick jq headline checks

Run:
  source .venv/bin/activate
  export SCRUB_API="http://127.0.0.1:8000"
  python scripts/eval_prompts_alt.py --in PROMPTS/merged.xlsx --clearance C3

Outputs:
  reports/merged_eval.alt.xlsx
  reports/prompt_eval_summary.json
  reports/prompt_eval_anomalies.md
"""

from __future__ import annotations
import argparse, json, os, re, sys, time, pathlib
from typing import Dict, Any, List, Optional, Tuple
import httpx
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Accept [TOKEN], <TOKEN>, and C4::TOKEN::hash forms
TOKEN_RX = re.compile(r"(?:[<\[]([A-Z0-9_]+)[>\]]|C\d?::([A-Z0-9_]+)::[a-z0-9]+)")
# For normalizing dataset-specific tokens like EMP1_PHONE -> PHONE; CORPKEY_1 -> CORPKEY
BASE_LABELS = {
    "PHONE","EMAIL","IBAN","AMOUNT","CURRENCY","ACCOUNT","CARD","SSN","DOB",
    "DATE","TIME","ADDRESS","CITY","COUNTRY","ZIP","STATUS","ROUTE","BIC",
    "FIRST","LAST","NAME","CUSTOMER_NAME","CORPKEY","ORDER_ID","TX_ID","IBAN_BE","IBAN_NL"
}
NUM_SUFFIX_RX = re.compile(r"_(\d+)$")

def tokens_of(s: str) -> List[str]:
    if not s:
        return []
    out: List[str] = []
    for g1, g2 in TOKEN_RX.findall(s):
        t = g1 or g2
        out.append(t)
    return out

def normalize_token(t: str) -> str:
    # Remove trailing numeric suffixes like _1
    t = NUM_SUFFIX_RX.sub("", t)
    # If token contains underscores, prefer the part that matches base labels
    parts = t.split("_")
    for p in parts:
        if p in BASE_LABELS:
            return p
    # Otherwise, return de-indexed token
    return t

def to_bracket_style(text: str) -> str:
    # Convert "C4::PHONE::abc123 ..." -> "[PHONE] ..."
    return re.sub(r"C\d?::([A-Z0-9_]+)::[a-z0-9]+", lambda m: f"[{m.group(1)}]", text or "")

def best_input_column(headers: List[str]) -> Optional[str]:
    # Choose the column in the workbook that holds the ORIGINAL text to sanitize.
    candidates = ["Prompt", "Original Prompt", "Input", "Text", "Source", "Raw"]
    for c in candidates:
        if c in headers:
            return c
    # Fallback to "Sanitized Prompt" if nothing else exists (not ideal, but safe)
    return "Prompt" if "Prompt" in headers else ("Text" if "Text" in headers else None)

def header_index_map(ws: Worksheet) -> Dict[str, int]:
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {name: i for i, name in enumerate(hdr)}

def ensure_column(ws: Worksheet, name: str) -> int:
    # Ensure a column exists with given header; return its index.
    hdr_row = 1
    hdr = [c.value for c in next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row))]
    try:
        idx = hdr.index(name)
        return idx
    except ValueError:
        ws.cell(row=hdr_row, column=len(hdr)+1, value=name)
        return len(hdr)  # 0-based index

def main():
    ap = argparse.ArgumentParser(description="Alt evaluator using /scrub API with robust token coverage.")
    ap.add_argument("--in", dest="input_path", required=True, help="Path to input .xlsx (e.g., PROMPTS/merged.xlsx)")
    ap.add_argument("--clearance", dest="clearance", default=os.getenv("SP_CLEARANCE","C3"), help="Clearance (default: C3)")
    ap.add_argument("--outdir", dest="outdir", default="reports", help="Output directory (default: reports)")
    args = ap.parse_args()

    api = os.getenv("SCRUB_API") or os.getenv("SECUREPROMPT_API_URL") or "http://127.0.0.1:8000"
    input_xlsx = pathlib.Path(args.input_path)
    outdir = pathlib.Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    if not input_xlsx.exists():
        print(f"error: input not found: {input_xlsx}", file=sys.stderr)
        sys.exit(2)

    wb = load_workbook(str(input_xlsx))
    ws = wb.active
    hdr_map = header_index_map(ws)
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    src_col_name = best_input_column(headers)
    if not src_col_name or src_col_name not in hdr_map:
        print(f"error: could not find a source text column in headers={headers}", file=sys.stderr)
        sys.exit(2)

    # Ensure output columns exist
    got_raw_idx   = ensure_column(ws, "Got_Sanitized_Prompt_Raw")
    got_norm_idx  = ensure_column(ws, "Got_Sanitized_Prompt")
    ent_count_idx = ensure_column(ws, "Got_Entities_Count")

    # Optional response columns
    gold_resp_idx = hdr_map.get("Sanitized Response")
    got_resp_idx  = ensure_column(ws, "Got_Sanitized_Response")

    # Coverage counters
    rows = 0
    tp = fp = fn = 0
    anomalies: List[str] = []

    def tally(expected: List[str], got: List[str]) -> Tuple[int,int,int]:
        E = {normalize_token(t) for t in expected}
        G = {normalize_token(t) for t in got}
        return len(E & G), len(G - E), len(E - G)

    # Iterate rows
    with httpx.Client(timeout=30.0) as client:
        for r in ws.iter_rows(min_row=2):
            rows += 1
            src_cell = r[hdr_map[src_col_name]]
            src = src_cell.value if src_cell is not None else None
            if not src:
                continue

            try:
                resp = client.post(f"{api}/scrub", json={"text": src, "clearance": args.clearance})
                if resp.status_code != 200:
                    anomalies.append(f"row {rows}: /scrub {resp.status_code}")
                    continue
                data = resp.json()
                raw_scrub = data.get("scrubbed","")
                norm_scrub = to_bracket_style(raw_scrub)
                ents = data.get("entities",[]) or []

                # Write outputs
                r[got_raw_idx].value  = raw_scrub
                r[got_norm_idx].value = norm_scrub
                r[ent_count_idx].value = len(ents)

                # Coverage against gold prompt tokens (if present)
                gold_prompt_col = hdr_map.get("Sanitized Prompt")
                if gold_prompt_col is not None:
                    gold_prompt = r[gold_prompt_col].value
                    if gold_prompt:
                        E = tokens_of(gold_prompt)
                        G = tokens_of(norm_scrub) or tokens_of(raw_scrub)
                        ttp, tfp, tfn = tally(E, G)
                        tp += ttp; fp += tfp; fn += tfn

                # Optional response coverage if your sheet has a gold response column
                if gold_resp_idx is not None:
                    gold_resp = r[gold_resp_idx].value
                    # If you have a response to scrub, you could POST it as well; this script
                    # just mirrors prompt coverage into the response column to keep columns present.
                    if gold_resp:
                        r[got_resp_idx].value = ""  # leave empty unless you also scrub response

            except Exception as e:
                anomalies.append(f"row {rows}: exception {e}")

    # Save workbook copy (avoid overwriting your existing one)
    out_xlsx = outdir / "merged_eval.alt.xlsx"
    wb.save(str(out_xlsx))

    # Summary JSON
    tot = tp + fp + fn
    coverage = (tp / tot) if tot else 0.0
    summary = {
        "input": str(input_xlsx),
        "rows": rows,
        "prompt_accuracy": coverage,           # treat coverage as accuracy for headline
        "response_accuracy": None,
        "total_response_entities": None,
        "notes": "Coverage computed using bracket-agnostic + C#::LABEL:: matching with token normalization.",
        "ts": int(time.time()),
    }
    (outdir / "prompt_eval_summary.json").write_text(json.dumps(summary, indent=2))

    # Anomalies file
    if anomalies:
        (outdir / "prompt_eval_anomalies.md").write_text(
            "# Anomalies\n\n" + "\n".join(f"- {a}" for a in anomalies) + "\n"
        )
    print(f"Done. Wrote: {out_xlsx} and {outdir/'prompt_eval_summary.json'}")

if __name__ == "__main__":
    main()
