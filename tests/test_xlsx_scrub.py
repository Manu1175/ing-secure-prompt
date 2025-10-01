from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from secureprompt.files.xlsx import scrub_workbook
from secureprompt.receipts.store import read_receipt


def test_xlsx_scrubbing_pipeline(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Email a@b.com"
    src_path = tmp_path / "example.xlsx"
    workbook.save(src_path)

    result = scrub_workbook(src_path, clearance="C3", filename="example.xlsx")

    redacted_path = Path(result["redacted_path"])
    assert redacted_path.exists()

    redacted_wb = load_workbook(redacted_path, read_only=True, data_only=True)
    try:
        redacted_value = redacted_wb["Sheet1"]["A1"].value
    finally:
        redacted_wb.close()

    assert redacted_value is not None
    assert "a@b.com" not in str(redacted_value)

    receipt = read_receipt(result["operation_id"])
    excel_meta = [entity.get("excel") for entity in receipt.get("entities", []) if entity.get("excel")]
    assert excel_meta, "Expected at least one entity with Excel metadata"
    assert excel_meta[0]["sheet"] == "Sheet1"
    assert excel_meta[0]["cell"] == "A1"
